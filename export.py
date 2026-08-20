"""
export.py
Port of build_export_workbook() from global.R: reproduces the original
Block Template.xlsx layout (merged section-header row + field-name row) with
every block's data, plus a "Submission Status" sheet.
"""
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import meta

SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HDR_FONT = Font(bold=True)
_thin = Side(style="thin")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_export_workbook(entries: pd.DataFrame) -> openpyxl.Workbook:
    all_meta = meta.field_meta  # already sorted by idx
    ncols = len(all_meta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Block Epi Info"

    sections = all_meta["section"].tolist()
    for j, (_, row) in enumerate(all_meta.iterrows(), start=1):
        c1 = ws.cell(row=1, column=j, value=row["section"])
        c1.fill, c1.font, c1.border, c1.alignment = SECTION_FILL, SECTION_FONT, BORDER, CENTER_WRAP
        c2 = ws.cell(row=2, column=j, value=row["field_name"])
        c2.fill, c2.font, c2.border, c2.alignment = HDR_FILL, HDR_FONT, BORDER, CENTER_WRAP
        ws.column_dimensions[get_column_letter(j)].width = 18

    # merge contiguous identical section cells in row 1
    i = 1
    while i <= ncols:
        j = i
        while j < ncols and sections[j - 1] == sections[j]:
            j += 1
        if j > i:
            ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=j)
        i = j + 1

    ws.freeze_panes = "E3"

    if not entries.empty:
        entries = entries.copy()
        loc_id_state = all_meta.loc[all_meta.idx == 2.0, "field_id"].iloc[0]
        loc_id_district = all_meta.loc[all_meta.idx == 3.0, "field_id"].iloc[0]
        loc_id_block = all_meta.loc[all_meta.idx == 4.0, "field_id"].iloc[0]
        entries[loc_id_state] = entries["state"]
        entries[loc_id_district] = entries["district"]
        entries[loc_id_block] = entries["block"]

        ordered_ids = all_meta["field_id"].tolist()
        for fid in ordered_ids:
            if fid not in entries.columns:
                entries[fid] = None
        export_df = entries[ordered_ids]

        for r, (_, erow) in enumerate(export_df.iterrows(), start=3):
            for c, fid in enumerate(ordered_ids, start=1):
                val = erow[fid]
                ws.cell(row=r, column=c, value=(None if pd.isna(val) else val))

    ws2 = wb.create_sheet("Submission Status")
    status_cols = [c for c in ["state", "district", "block", "status", "submitted_by", "updated_at"]
                   if c in entries.columns]
    ws2.append(status_cols)
    if not entries.empty:
        for _, srow in entries[status_cols].iterrows():
            ws2.append([None if pd.isna(v) else v for v in srow.tolist()])

    return wb
